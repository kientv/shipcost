#!/usr/bin/env python3
"""
Extract all pricing data from Excel to JSON
"""
import openpyxl
import json
import re

def clean(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    return s if s else None

def to_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(',', '.').replace(' kg', '').replace('kg', '').strip()
    try:
        return float(s)
    except:
        return None

wb = openpyxl.load_workbook('BẢNG GIÁ NEWPOST PUPLIC KHÁCH HÀNG.xlsx', data_only=True)

output = {}

# ===== 1. CTU (Chuyên Tuyến ỦY Thác) =====
ws = wb['Chuyên Tuyến Uỷ Thác (CTU)']
ctu = {"countries": [], "timeline": {}, "rates": []}
# Row 10: countries
for row in ws.iter_rows(min_row=10, max_row=10, values_only=True):
    ctu["countries"] = [clean(c) for c in row[1:] if clean(c)]
# Row 11: timeline
for row in ws.iter_rows(min_row=11, max_row=11, values_only=True):
    tls = [clean(c) for c in row[1:] if clean(c)]
    for i, c in enumerate(ctu["countries"]):
        if i < len(tls):
            ctu["timeline"][c] = tls[i]
# Rows 12+: rates
for row in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if vals[0] and to_float(vals[0]) is not None:
        w = to_float(vals[0])
        rates = {}
        for i, c in enumerate(ctu["countries"]):
            if i+1 < len(vals) and vals[i+1] is not None:
                rates[c] = to_float(vals[i+1])
        if rates:
            ctu["rates"].append({"weight": w, "rates": rates})
output["ctu"] = ctu

# ===== 2. CTQ (Chuyên Tuyến) =====
ws = wb['Chuyên Tuyến (CTQ)']
ctq = {"document": {}, "non_document": {}}
# Row 8: headers
for row in ws.iter_rows(min_row=8, max_row=8, values_only=True):
    headers = [clean(c) for c in row]
countries = [c for c in headers[1:] if c]
in_non_doc = False
for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if not any(v for v in vals):
        continue
    if "For DOC" in str(vals) or "For DOX" in str(vals):
        in_non_doc = True
        continue
    if vals[0] and vals[0] not in ['Weight (Kg)']:
        w = to_float(vals[0])
        if w is not None:
            target = ctq["non_document"] if in_non_doc else ctq["document"]
            if w not in target:
                target[w] = {}
            for i, c in enumerate(countries):
                if i+1 < len(vals) and vals[i+1] is not None:
                    target[w][c] = to_float(vals[i+1])
output["ctq"] = ctq

# ===== 3. DHL SING =====
ws = wb['DHL SING']
dhl_sing = {"doc": {}, "non_doc": {}}
in_non_doc = False
zone_headers = None
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if not any(v for v in vals):
        continue
    if "Non-Documents" in str(vals):
        in_non_doc = True
        zone_headers = None
        continue
    if vals[0] and "Weight" in str(vals[0]):
        zone_headers = [c for c in vals[1:] if c]
        continue
    if vals[0] and zone_headers:
        w = to_float(vals[0])
        if w is not None:
            target = dhl_sing["non_doc"] if in_non_doc else dhl_sing["doc"]
            if w not in target:
                target[w] = {}
            for i, z in enumerate(zone_headers):
                if i+1 < len(vals) and vals[i+1] is not None:
                    target[w][z] = to_float(vals[i+1])
output["dhl_sing"] = dhl_sing

# ===== 4. DHL VIET NAM (DHLV) =====
ws = wb['DHL VIET NAM (DHLV)']
dhlv = {"doc": {}, "non_doc": {}, "transit": {}}
in_non_doc = False
zone_headers = None
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if not any(v for v in vals):
        continue
    if "Transit time" in str(vals):
        for i, z in enumerate([c for c in vals[1:] if c]):
            dhlv["transit"][z] = vals[i+1] if i+1 < len(vals) else ""
        continue
    if "For DOX" in str(vals) or "For DOC" in str(vals):
        in_non_doc = True
        zone_headers = None
        continue
    if "Outbound" in str(vals) or "Package Express" in str(vals):
        continue
    if vals[0] and "Weight" in str(vals[0]):
        zone_headers = [c for c in vals[1:] if c]
        continue
    if vals[0] and zone_headers:
        w = to_float(vals[0])
        if w is not None:
            target = dhlv["non_doc"] if in_non_doc else dhlv["doc"]
            if w not in target:
                target[w] = {}
            for i, z in enumerate(zone_headers):
                if i+1 < len(vals) and vals[i+1] is not None:
                    target[w][z] = to_float(vals[i+1])
output["dhl_vietnam"] = dhlv

# ===== 5. ZONE DHL SING =====
ws = wb['ZONE DHL SING']
zone_dhl_sing = {}
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    for i in range(0, len(vals), 2):
        if i+1 < len(vals) and vals[i] and vals[i+1]:
            try:
                zone_dhl_sing[vals[i]] = int(vals[i+1])
            except:
                pass
output["zone_dhl_sing"] = zone_dhl_sing

# ===== 6. ZONE DHL (DHLV) =====
ws = wb['ZONE DHL (DHLV)']
zone_dhlv = {}
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    for i in range(0, len(vals), 2):
        if i+1 < len(vals) and vals[i] and vals[i+1]:
            try:
                zone_dhlv[vals[i]] = int(vals[i+1])
            except:
                pass
output["zone_dhl_vietnam"] = zone_dhlv

# ===== 7. PHỤ PHÍ DHL VN =====
ws = wb['PHỤ PHÍ DHL VN']
surcharge_dhlv = []
for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if len(vals) >= 3 and vals[0] and vals[1]:
        surcharge_dhlv.append({
            "type": vals[0],
            "rate": vals[1],
            "note": vals[2] if len(vals) > 2 else ""
        })
output["surcharge_dhl_vietnam"] = surcharge_dhlv

# ===== 8. UPS SAVER =====
ws = wb['UPS SAVER']
ups_saver = {"envelope": {}, "document": {}, "parcel": {}}
zone_headers = None
current_type = None
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if not any(v for v in vals):
        continue
    # Type header rows can also contain data (e.g., "UPS Envelope", 0.5, ...)
    is_type_header = vals[0] in ['UPS Envelope', 'Tài liệu', 'Hàng hóa']
    if is_type_header:
        type_map = {
            'UPS Envelope': 'envelope',
            'Tài liệu': 'document',
            'Hàng hóa': 'parcel'
        }
        current_type = type_map[vals[0]]
        # Also extract data from this row (weight in column 1)
        if zone_headers and vals[1] is not None:
            w = to_float(vals[1])
            if w is not None:
                if w not in ups_saver[current_type]:
                    ups_saver[current_type][w] = {}
                for i, z in enumerate(zone_headers):
                    col_idx = 2 + i
                    if col_idx < len(vals) and vals[col_idx] is not None:
                        ups_saver[current_type][w][z] = to_float(vals[col_idx])
        continue
    if zone_headers is None and len(vals) > 1 and vals[1] and "KL" in str(vals[1]):
        zone_headers = [v for v in vals[2:] if v]
        continue
    if current_type and zone_headers:
        # Weight is in column 1 (index 1) for data rows (column 0 is None/empty)
        if vals[0] is None and vals[1] is not None:
            w = to_float(vals[1])
            if w is not None:
                if w not in ups_saver[current_type]:
                    ups_saver[current_type][w] = {}
                for i, z in enumerate(zone_headers):
                    col_idx = 2 + i
                    if col_idx < len(vals) and vals[col_idx] is not None:
                        ups_saver[current_type][w][z] = to_float(vals[col_idx])
output["ups_saver"] = ups_saver

# ===== 9. UPS ZONE =====
ws = wb['UPS ZONE']
ups_zone = {}
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if len(vals) >= 2 and vals[0] and vals[1] and vals[0] != 'Rate Guide Country Name':
        ups_zone[vals[0]] = vals[1]
output["ups_zone"] = ups_zone

# ===== 10. UPS PHỤ PHÍ =====
ws = wb['UPS PHỤ PHÍ']
ups_surcharge = []
for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if len(vals) >= 4 and vals[0] and vals[1]:
        ups_surcharge.append({
            "id": vals[0],
            "type": vals[1],
            "definition": vals[2],
            "rate": vals[3]
        })
output["ups_surcharge"] = ups_surcharge

# ===== 11. NP_FEDEX =====
ws = wb['NP_FEDEX']
fedex = {"envelope": {}, "pak": {}, "ip": {}}
zone_headers = None
current_service = None
for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if not any(v for v in vals):
        continue
    if vals[0] == 'ZONE':
        zone_headers = [str(c) for c in vals[2:] if c]
        continue
    if vals[0] in ['FedEx Envelope2', 'FedEx Pak2', 'International Priority (IP)']:
        svc_map = {
            'FedEx Envelope2': 'envelope',
            'FedEx Pak2': 'pak',
            'International Priority (IP)': 'ip'
        }
        current_service = svc_map[vals[0]]
        # First row of service has weight in column 1
        w = to_float(vals[1])
        if w is not None:
            if w not in fedex[current_service]:
                fedex[current_service][w] = {}
            for i, z in enumerate(zone_headers):
                if i+2 < len(vals) and vals[i+2] is not None:
                    fedex[current_service][w][z] = to_float(vals[i+2])
        continue
    if vals[0] == '' and vals[1] and zone_headers and current_service:
        # Continuation rows have empty column 0, weight in column 1
        w = to_float(vals[1])
        if w is not None:
            if w not in fedex[current_service]:
                fedex[current_service][w] = {}
            for i, z in enumerate(zone_headers):
                if i+2 < len(vals) and vals[i+2] is not None:
                    fedex[current_service][w][z] = to_float(vals[i+2])
output["fedex"] = fedex

# ===== 12. ZONE FEDEX =====
ws = wb['ZONE FEDEX']
zone_fedex = {}
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    # Data is in columns 1,2 then 5,6 then 9,10 (0-indexed)
    for col_start in [1, 5, 9]:
        if col_start + 1 < len(vals) and vals[col_start] and vals[col_start + 1]:
            try:
                zone_fedex[vals[col_start]] = int(vals[col_start + 1])
            except:
                pass
output["zone_fedex"] = zone_fedex

# ===== 13. PHỤ PHÍ HQ =====
ws = wb['Phụ phí HQ']
surcharge_hq = []
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
    vals = [clean(c) for c in row]
    if len(vals) >= 4 and vals[1]:
        surcharge_hq.append({
            "id": vals[0],
            "item": vals[1],
            "rate": vals[2],
            "note": vals[3] if len(vals) > 3 else ""
        })
output["surcharge_hq"] = surcharge_hq

# Save
with open('pricing_data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print("Saved pricing_data.json")
for k, v in output.items():
    if isinstance(v, dict):
        print(f"  {k}: {len(v)} keys")
    elif isinstance(v, list):
        print(f"  {k}: {len(v)} items")
    else:
        print(f"  {k}: {type(v)}")